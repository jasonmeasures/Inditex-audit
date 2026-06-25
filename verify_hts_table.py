#!/usr/bin/env python3
"""Verify the HTS Classification Table XLSX parses like the dashboard parser."""

from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_PATH = Path(__file__).resolve().parent / "HTS_Classification_Table (1).xlsx"


def norm(s: str) -> str:
    return re.sub(r"[.\s]", "", str(s or ""))


def parse_hts_table(path: Path) -> dict[str, dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(str(v or "").strip() == "HTS No." for v in vals):
            header_row = r
            headers = [str(v or "").strip() for v in vals]
            break
    if header_row is None:
        raise SystemExit("FAIL: could not find header row with 'HTS No.'")

    required = ["HTS No.", "C1 Ad Valorem formula", "C1 Ad Valorem", "Description", "Updated"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise SystemExit(f"FAIL: missing columns: {missing}")

    idx = {h: headers.index(h) for h in required}
    spec_idx = headers.index("C1 Rate Specific formula") if "C1 Rate Specific formula" in headers else None
    lookup: dict[str, dict] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        hts_raw = ws.cell(r, idx["HTS No."] + 1).value
        if not hts_raw:
            continue
        hts = norm(hts_raw)
        if len(hts) != 10:
            continue
        formula = ws.cell(r, idx["C1 Ad Valorem formula"] + 1).value
        if isinstance(formula, (int, float)) and formula == formula:
            rate_pct = float(formula) * 100
        else:
            rate_pct = float(str(ws.cell(r, idx["C1 Ad Valorem"] + 1).value or "0"))
        has_specific = False
        if spec_idx is not None:
            spec = ws.cell(r, spec_idx + 1).value
            has_specific = isinstance(spec, (int, float)) and bool(spec) and spec > 0
        desc = str(ws.cell(r, idx["Description"] + 1).value or "")
        upd = ws.cell(r, idx["Updated"] + 1).value
        upd_ms = upd.timestamp() * 1000 if isinstance(upd, datetime) else 0
        existing = lookup.get(hts)
        if not existing or upd_ms > existing["_updMs"]:
            lookup[hts] = {
                "rate_pct": rate_pct,
                "has_specific": has_specific,
                "description": desc[:60],
                "_updMs": upd_ms,
            }
    return lookup


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH
    if not path.exists():
        print(f"FAIL: file not found: {path}")
        return 1

    lookup = parse_hts_table(path)
    if len(lookup) < 1000:
        print(f"FAIL: only {len(lookup)} codes parsed — expected thousands")
        return 1

    print(f"PASS: {len(lookup):,} HTS codes from {path.name}")
    samples = ["6109100012", "6110201020", "6204628056"]
    for code in samples:
        entry = lookup.get(code)
        if entry:
            print(f"  {code}: {entry['rate_pct']}% — {entry['description']}")
        else:
            print(f"  {code}: (not in table)")
    rates = [v["rate_pct"] for v in lookup.values()]
    print(f"  rate range: {min(rates):.4f}% – {max(rates):.4f}%")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
